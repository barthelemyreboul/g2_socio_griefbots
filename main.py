import os
import re
from collections import Counter
from datetime import datetime

import nltk
import praw
from dotenv import load_dotenv
from nltk.corpus import stopwords
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from praw.models import Subreddit

load_dotenv()
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")


def get_words_lists() -> tuple[list[str], list[str]]:
    positive_words = []
    negative_words = []
    with open("words_lists/positive_words.txt", "r") as pos_file:
        positive_words = [
            line.strip()
            for line in pos_file
            if line.strip() and not line.startswith(";")
        ]
    with open("words_lists/negative_words.txt", "r") as neg_file:
        negative_words = [
            line.strip()
            for line in neg_file
            if line.strip() and not line.startswith(";")
        ]
    return positive_words, negative_words


def analyze_post_interactions(subreddit: Subreddit, type: str, limit: int = 5) -> None:
    """Analyze interactions on the most popular posts in a subreddit.
    Args:
        subreddit (Subreddit): The subreddit to analyze.
        type (str): Type of posts to analyze ("hot", "top", or "new").
        limit (int): Number of top posts to analyze.
    """
    # Select posts based on the specified type
    if type == "hot":
        print(
            f"\n=== Analysing top {limit} hot posts in r/{subreddit.display_name} ===\n"
        )
        reddit_iterator = subreddit.hot(limit=limit)
    elif type == "top":
        print(
            f"\n=== Analysing {limit} toppest posts in r/{subreddit.display_name} ===\n"
        )
        reddit_iterator = subreddit.top(limit=limit)
    else:
        print(
            f"\n=== Analysing {limit} newest posts in r/{subreddit.display_name} ===\n"
        )
        reddit_iterator = subreddit.new(limit=limit)

    # Load positive and negative words
    positive_words, negative_words = get_words_lists()

    # Getting common English stopwords
    nltk.download("stopwords")
    common_stopwords = set(stopwords.words("english"))

    posts_stats = []

    # Analyze each post
    for post in reddit_iterator:
        # Fetch all comments
        post.comments.replace_more(limit=0)
        comments = post.comments.list()
        if post.treatment_tags:
            print(f"Skipping post '{post.title}' (advertisement detected).")
            continue

        # Basic statistics on comments
        nb_comments = len(comments)
        avg_score = (
            sum(comment.score for comment in comments) / nb_comments
            if nb_comments > 0
            else 0
        )
        avg_length = (
            sum(len(comment.body) for comment in comments) / nb_comments
            if nb_comments > 0
            else 0
        )

        # Sentiment analysis on comments
        pos_count = sum(
            any(w in comment.body.lower() for w in positive_words)
            for comment in comments
        )
        neg_count = sum(
            any(w in comment.body.lower() for w in negative_words)
            for comment in comments
        )

        # Most frequent words excluding short/common words on comments
        words = [
            w.lower()
            for comment in comments
            for w in comment.body.split()
            if len(w) > 3 and w.isalpha() and w.lower() not in common_stopwords
        ]
        common_words = Counter(words).most_common(5)

        posts_stats.append(
            {
                "title": post.title[:60],  # tronqué pour lisibilité
                "score": post.score,
                "positive": pos_count,
                "negative": neg_count,
            }
        )

        # Print analysis results
        print(f"Title: {post.title}")
        print(f"Author: {post.author}")
        print(f"Score (likes): {post.score}")
        print(f"Number of comments: {nb_comments}")
        print(f"Average comment score: {avg_score:.2f}")
        print(f"Average comment length: {avg_length:.1f} characters")
        print(f"Estimated sentiment: +{pos_count} / -{neg_count}")
        print(f"Most frequent words: {common_words}")
        print("-" * 100)
    # return plot_sentiment_bars(posts_stats)


def contains_keyword(text: str, keywords: list[str]) -> bool:
    """
    Checks if any of the keywords are present in the given text (whole words only).
    Args:
        text (str): The text to search within.
        keywords (list[str]): List of keywords to search for.
    Returns:
        bool: True if any keyword is found, False otherwise.
    """
    if not text or not keywords:
        return False
    for kw in keywords:
        pattern = rf"\b{re.escape(kw)}\b"
        if re.search(pattern, text, flags=re.IGNORECASE):
            return True
    return False


def extract_post_data(
    subreddit: Subreddit,
    limit: int = 5,
    start_date: datetime | None = None,
    end_date: datetime | None = None,
    key_words: list[str] | None = None,
) -> list[dict]:
    """
    Extracts posts and comments containing given keywords or within a date range.
    Args:
        subreddit (Subreddit): The subreddit to extract data from.
        limit (int): Maximum number of posts and comments to extract.
        start_date (datetime | None): Start date for filtering posts/comments.
        end_date (datetime | None): End date for filtering posts/comments.
        key_words (list[str] | None): List of keywords to filter posts/comments.
    Returns:
        list[dict]: List of extracted post and comment data.

    """
    all_data = []
    posts = subreddit.top(limit=None)
    post_count = 0
    scrapped = 0

    for post in posts:
        scrapped += 1
        if post_count >= limit:
            break

        # Filter post by date
        post_datetime = datetime.utcfromtimestamp(post.created_utc)
        if start_date and post_datetime < start_date:
            continue
        if end_date and post_datetime > end_date:
            continue

        # Check post keywords
        post_content = getattr(post, "selftext", "") or ""
        post_has_keyword = (
            contains_keyword(post_content, key_words) if key_words else True
        )

        # Build post data
        post_id = f"RD-{post_count + 1:02d}"
        post_data = {
            "ID": post_id,
            "Reddit URL": f"https://www.reddit.com{post.permalink}"
            if hasattr(post, "permalink")
            else None,
            "Thread Title": getattr(post, "title", None),
            "Subreddit": getattr(subreddit, "display_name", None),
            "Content": post_content.replace("\n", " ").strip() or None,
            "Category": "post",
            "Author": ("u/" + str(post.author))
            if getattr(post, "author", None)
            else None,
            "Date Posted": post_datetime.strftime("%d/%m/%Y"),
            "Number of comments": getattr(post, "num_comments", 0),
            "Upvotes": getattr(post, "score", 0),
            "Keywords": None,
            "Sentiment": None,
            "Tag": getattr(post, "link_flair_text", None),
            "Date added": datetime.utcnow().strftime("%d/%m/%Y"),
            "Who added": "automated_script",
        }

        # Load comments
        post.comments.replace_more(limit=0)
        comments = post.comments.list()
        comment_count = 0
        relevant_comments = []

        for comment in comments:
            scrapped += 1
            if comment_count >= limit:
                break
            if not hasattr(comment, "body") or not hasattr(comment, "created_utc"):
                continue

            comment_text = getattr(comment, "body", "")
            comment_datetime = datetime.utcfromtimestamp(comment.created_utc)

            # Filter comment by date
            if start_date and comment_datetime < start_date:
                continue
            if end_date and comment_datetime > end_date:
                continue

            # Check comment for keywords
            comment_has_keyword = (
                contains_keyword(comment_text, key_words) if key_words else True
            )
            if not comment_has_keyword:
                continue

            comment_id = f"{post_id}-{comment_count + 1:02d}"
            comment_data = {
                "ID": comment_id,
                "Reddit URL": f"https://www.reddit.com{getattr(comment, 'permalink', '')}"
                if hasattr(comment, "permalink")
                else None,
                "Thread Title": getattr(post, "title", None),
                "Subreddit": "r/" + getattr(subreddit, "display_name", None),
                "Content": comment_text.replace("\n", " ").strip() or None,
                "Category": "comment",
                "Author": (
                    "u/" + (str(comment.author))
                    if getattr(comment, "author", None)
                    else ""
                ),
                "Date Posted": comment_datetime.strftime("%d/%m/%Y"),
                "Number of comments": None,
                "Upvotes": getattr(comment, "score", 0),
                "Keywords": None,
                "Sentiment": None,
                "Tag": getattr(post, "link_flair_text", None),
                "Date added": datetime.utcnow().strftime("%d/%m/%Y"),
                "Who added": "automated_script",
            }
            relevant_comments.append(comment_data)
            comment_count += 1

        # Add post and its relevant comments if criteria met
        if post_has_keyword or relevant_comments:
            all_data.append(post_data)
            all_data.extend(relevant_comments)
            post_count += 1
            print(f"Added post {post_id} ({len(relevant_comments)} comments)")

    print(f"✅ Total relevant items: {len(all_data)} / Scrapped: {scrapped}")
    return all_data


def save_subreddits_to_excel(
    reddit: Subreddit,
    subreddit_names: list,
    output_path: str = "reddit_threads.xlsx",
    limit: int = 5,
    start_date: datetime | None = None,
    end_date: datetime | None = None,
    key_words: list[str] | None = None,
) -> None:
    """Saves extracted subreddit data to an Excel file.
    Args:
        reddit (Subreddit): The Reddit instance.
        subreddit_names (list): List of subreddit names to extract data from.
        output_path (str): Path to save the Excel file.
        limit (int): Maximum number of posts and comments to extract per subreddit.
        start_date (datetime | None): Start date for filtering posts/comments.
        end_date (datetime | None): End date for filtering posts/comments.
    """
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    columns = [
        "ID",
        "Reddit URL",
        "Thread Title",
        "Subreddit",
        "Content",
        "Category",
        "Author",
        "Date Posted",
        "Number of comments",
        "Upvotes",
        "Keywords",
        "Sentiment",
        "Tag",
        "Date added",
        "Who added",
    ]

    for sub_name in subreddit_names:
        subreddit = reddit.subreddit(sub_name)
        all_data = extract_post_data(
            subreddit,
            limit=limit,
            start_date=start_date,
            end_date=end_date,
            key_words=key_words,
        )

        ws = wb.create_sheet(title=sub_name[:31])
        # Headers
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # columns data
        current_row = 2
        for item in all_data:
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=current_row, column=col_idx, value=item.get(col_name, None))
            current_row += 1

        # Size columns
        for col_idx, col_name in enumerate(columns, start=1):
            max_length = max(
                (
                    len(str(ws.cell(row=row, column=col_idx).value))
                    for row in range(1, ws.max_row + 1)
                    if ws.cell(row=row, column=col_idx).value
                ),
                default=0,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_length + 2, 60
            )

    wb.save(output_path)


if __name__ == "__main__":
    reddit = praw.Reddit(
        client_id=CLIENT_ID,
        client_secret=CLIENT_SECRET,
        user_agent="python:projectdecember.script:v1.0 (by u/sk00bew)",
    )

    begin = datetime(2020, 1, 1)
    end = None
    # "GriefSupport", Futurology
    threads = ["Chatbots", "artificial", "ArtificialInteligence"]

    save_subreddits_to_excel(
        reddit,
        threads,
        output_path="reddit_threads.xlsx",
        limit=30,
        start_date=begin,
        end_date=end,
        key_words=key_words_1,
    )
