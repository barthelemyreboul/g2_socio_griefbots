from datetime import datetime
import praw
from praw import Reddit
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from config import key_words_1,key_words_2,key_words_3
from utils import extract_post_data, CLIENT_ID, CLIENT_SECRET


def save_subreddits_to_excel(
    reddit: Reddit,
    subreddit_names: list[str],
    limit: int = 5,
    start_date: datetime | None = None,
    end_date: datetime | None = None,
    key_words: list[str] | None = None,
) -> None:
    """Saves extracted subreddit data to an Excel file.
    Args:
        reddit (Reddit): The Reddit instance.
        subreddit_names (list): List of subreddit names to extract data from.
        output_path (str): Path to save the Excel file.
        limit (int): Maximum number of posts and comments to extract per subreddit.
        start_date (datetime | None): Start date for filtering posts/comments.
        end_date (datetime | None): End date for filtering posts/comments.
        key_words (list[str] | None): List of keywords to filter posts/comments.
    """
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    columns = [
        "id",
        "Reddit_URL",
        "Thread_Title",
        "Subreddit",
        "Content",
        "Category",
        "Author",
        "Date_Posted",
        "Number_of_comments",
        "Upvotes",
        "Keywords",
        "Sentiment",
        "Tag",
        "Date_added",
        "Who_added",
    ]

    for sub_name in subreddit_names:
        subreddit = reddit.subreddit(sub_name)
        all_data = extract_post_data(
            subreddit,
            limit=limit,
            start_date=start_date,
            end_date=end_date,
            key_words=key_words,
        )

        ws = wb.create_sheet(title=sub_name[:31])
        # Headers
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # columns data
        current_row = 2
        for item in all_data:
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=current_row, column=col_idx, value=getattr(item,col_name, None))
            current_row += 1

        # Size columns
        for col_idx, col_name in enumerate(columns, start=1):
            max_length = max(
                (
                    len(str(ws.cell(row=row, column=col_idx).value))
                    for row in range(1, ws.max_row + 1)
                    if ws.cell(row=row, column=col_idx).value
                ),
                default=0,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_length + 2, 60
            )

    wb.save(f"reddit_threads_1.xlsx")


if __name__ == "__main__":
    reddit = praw.Reddit(
        client_id=CLIENT_ID,
        client_secret=CLIENT_SECRET,
        user_agent="u/sk00bew",
    )

    begin = datetime(2020, 1, 1)
    end = None
    # "GriefSupport", Futurology
    threads = ["Chatbots", "artificial", "ArtificialInteligence"]


    save_subreddits_to_excel(
        reddit=reddit,
        subreddit_names=threads,
        limit=10,
        start_date=begin,
        end_date=end,
        key_words=key_words_2
    )
